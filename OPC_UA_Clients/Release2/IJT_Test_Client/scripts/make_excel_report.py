#!/usr/bin/env python3
"""
make_excel_report.py — Convert JUnit XML test results into a rich Excel workbook.

Reads:  test-results/pytest.xml   (or path given via --xml=FILE)
Writes: test-results/report.xlsx  (or path given via --out=FILE)

Sheets produced:
  Conformance Overview        — banner, run metadata, KPI grid mirroring the markdown report
  Test Outcome Counts         — headline counts (passed / failed / skipped / xfailed) by test area
  IJT Facet Breakdown         — IJT facet coverage, when CU JSON is present
  Conformance Unit Details    — one row per conformance unit, when CU JSON is present
  Profile Coverage Comparison — IJT coverage overview, when CU JSON is present
  All Test Cases              — every test: name, file, status, duration, skip/fail reason
  Test Failures               — only failed tests with full message
  Skipped Test Cases          — only skipped tests with category and reason
  Expected Failures           — xfailed and xpassed tests with reason

Usage:
  python scripts/make_excel_report.py
  python scripts/make_excel_report.py --xml test-results/pytest.xml --out test-results/report.xlsx
  python scripts/make_excel_report.py --help
"""

from __future__ import annotations

import argparse
import ast
import json
import os
import platform
import re
import sys
import xml.etree.ElementTree as ET  # nosec B405
from collections import Counter
from dataclasses import dataclass
from datetime import datetime, timezone
from importlib import metadata
from pathlib import Path
from typing import Any

# Bandit B405/B314 suppressions are limited to trusted JUnit XML from pytest.

try:
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERROR: openpyxl is required.  Run: pip install openpyxl", file=sys.stderr)
    sys.exit(1)

try:
    import yaml  # type: ignore[import-untyped]
except ImportError:
    yaml = None  # type: ignore[assignment]

# ── Colour palette ────────────────────────────────────────────────────────────
_GREEN = "FF92D050"  # passed
_RED = "FFFF0000"  # failed
_YELLOW = "FFFFFF00"  # skipped
_ORANGE = "FFFFC000"  # xfailed / xpassed
_BLUE = "FF9DC3E6"  # header rows
_GRAY = "FFF2F2F2"  # alternating row background
_DARK_GRAY = "FFD9E1F2"
_WHITE = "FFFFFFFF"
_LIGHT_GREEN = "FFE2F0D9"
_LIGHT_GREEN_NOTE = "FFEAF7E8"
_LIGHT_RED = "FFFFE5E5"
_LIGHT_YELLOW = "FFFFF2CC"
_LIGHT_ORANGE = "FFFCE4D6"
_LIGHT_BLUE = "FFDDEBF7"

_STATUS_COLOUR = {
    "passed": _GREEN,
    "failed": _RED,
    "skipped": _YELLOW,
    "xfailed": _ORANGE,
    "xpassed": _ORANGE,
    "error": _RED,
}

_CU_STATUS_COLOUR = {
    "supported": _LIGHT_GREEN,
    "partial": _LIGHT_GREEN_NOTE,
    "action_needed": _LIGHT_RED,
    "passed": _LIGHT_GREEN,
    "failed": _LIGHT_RED,
    "error": _LIGHT_RED,
    "not_supported": _GRAY,
    "blocked": _LIGHT_ORANGE,
    "untested": _GRAY,
    "skipped": _LIGHT_YELLOW,
    "unknown": _WHITE,
}

_PROJECT_ROOT = Path(__file__).resolve().parents[1]
_PROFILES_DIR = _PROJECT_ROOT / "profiles"
_DEFAULT_CU_JSON = _PROJECT_ROOT / "test-results" / "cu-compliance-report.json"
_DEFAULT_BASELINE_JSON = _PROJECT_ROOT / "test-results" / "report-baseline.json"
_CU_COMPLIANCE_KEYS = {"supported", "partial", "not_supported", "blocked", "action_needed", "untested"}
_FINDING_OUTCOMES = {"partial", "not_supported", "blocked", "action_needed"}
_NOT_APPLICABLE = "Not Applicable"
if str(_PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(_PROJECT_ROOT))

# Shared report logic lives in helpers/*.py.
# Markdown and Excel generators must use the same helpers to stay in sync.
from helpers.cu_registry import cu_display_name  # noqa: E402
from helpers.git_info import short_git_sha as _short_git_sha  # noqa: E402
from helpers.report_scoring import (
    ACTION_ITEM_LABEL_ORDER as _ACTION_ITEM_LABEL_ORDER,
)
from helpers.report_scoring import (
    CAPABILITY_NOTE_LABEL_ORDER as _CAPABILITY_NOTE_LABEL_ORDER,
)
from helpers.report_scoring import (
    CAPABILITY_SUPPORT_ICONS as _CAPABILITY_SUPPORT_ICONS,
)
from helpers.report_scoring import (
    STATUS_ORDER as _STATUS_ORDER,
)
from helpers.report_scoring import (
    action_items_context as _action_items_context,
)
from helpers.report_scoring import (
    conformance_score as _conformance_score,
)
from helpers.report_scoring import (
    format_pct as _fmt_pct,
)
from helpers.report_scoring import (
    format_status_counts as _format_status_counts,
)
from helpers.report_scoring import (
    informational_notes_context as _informational_notes_context,
)
from helpers.report_scoring import (
    is_healthy as _is_healthy,
)
from helpers.report_scoring import (
    outcome_label as _outcome_label,
)
from helpers.report_scoring import (
    pct_value as _pct_value,
)
from helpers.report_scoring import (
    status_color_excel as _status_color_excel,
)
from helpers.report_scoring import (
    status_count_key as _status_count_key,
)
from helpers.report_scoring import (
    status_for as _status_for,
)

# ── Data model ────────────────────────────────────────────────────────────────


@dataclass
class TestCase:
    classname: str
    name: str
    file: str
    status: str  # passed | failed | skipped | xfailed | xpassed | error
    duration: float
    message: str = ""  # skip reason, failure message, or xfail reason

    @property
    def area(self) -> str:
        """Top-level directory of the test file."""
        parts = Path(self.file).parts
        return parts[0] if parts else "unknown"

    @property
    def short_name(self) -> str:
        return self.name


@dataclass(frozen=True)
class FacetInfo:
    key: str
    display_name: str
    description: str
    spec_section: str
    kind: str
    conformance_units: list[str]


@dataclass(frozen=True)
class ProfileInfo:
    key: str
    name: str
    description: str
    facets: list[str]


@dataclass(frozen=True)
class CapabilitiesInfo:
    server_name: str
    active_profile: str
    supported_facets: list[str]
    overrides: dict[str, str]


# ── XML parsing ───────────────────────────────────────────────────────────────


def _text(elem: ET.Element | None) -> str:
    if elem is None:
        return ""
    parts = []
    if elem.text:
        parts.append(elem.text.strip())
    for child in elem:
        if child.text:
            parts.append(child.text.strip())
        if child.tail:
            parts.append(child.tail.strip())
    return "\n".join(p for p in parts if p)


def parse_junit_xml(path: Path) -> list[TestCase]:
    tree = ET.parse(path)  # nosec B314
    root = tree.getroot()

    # Handle both <testsuites><testsuite>... and bare <testsuite>...
    suites = root.findall(".//testsuite") if root.tag == "testsuites" else [root]

    cases: list[TestCase] = []
    for suite in suites:
        for tc in suite.findall("testcase"):
            classname = tc.get("classname", "")
            name = tc.get("name", "")
            file = tc.get("file", classname.replace(".", "/") + ".py")
            duration = float(tc.get("time", "0") or "0")

            failure = tc.find("failure")
            error = tc.find("error")
            skip = tc.find("skipped")

            if failure is not None:
                status = "failed"
                message = (failure.get("message") or "") + "\n" + (failure.text or "")
            elif error is not None:
                status = "error"
                message = (error.get("message") or "") + "\n" + (error.text or "")
            elif skip is not None:
                status = "skipped"
                message = skip.get("message") or _text(skip)
                # pytest encodes xfail as <skipped type="pytest.xfail" ...>
                # Fallback: some JUnit emitters prefix the message with "xfail" instead
                if (skip.get("type") or "").lower() == "pytest.xfail" or message.lower().startswith("xfail"):
                    status = "xfailed"
            else:
                status = "passed"
                message = ""

            cases.append(
                TestCase(
                    classname=classname,
                    name=name,
                    file=file,
                    status=status,
                    duration=duration,
                    message=message.strip(),
                )
            )

    return cases


# ── CU / profile metadata parsing ─────────────────────────────────────────────


def _load_json(path: Path | None) -> dict[str, Any] | None:
    if path is None or not path.exists():
        return None
    return json.loads(path.read_text(encoding="utf-8"))


def _load_yaml(path: Path) -> dict[str, Any]:
    if yaml is None or not path.exists():
        return {}
    loaded = yaml.safe_load(path.read_text(encoding="utf-8")) or {}
    return loaded if isinstance(loaded, dict) else {}


def _load_facets() -> dict[str, FacetInfo]:
    raw = _load_yaml(_PROFILES_DIR / "facets.yaml")
    facets: dict[str, FacetInfo] = {}
    for key, data in (raw.get("facets") or {}).items():
        if not isinstance(data, dict):
            continue
        units = [str(cu) for cu in data.get("conformance_units", [])]
        facets[str(key)] = FacetInfo(
            key=str(key),
            display_name=str(data.get("display_name") or _title_from_key(str(key))),
            description=str(data.get("description") or "").strip(),
            spec_section=str(data.get("spec_section") or ""),
            kind=str(data.get("kind") or "facet"),
            conformance_units=units,
        )
    return facets


def _load_profiles() -> dict[str, ProfileInfo]:
    profiles: dict[str, ProfileInfo] = {}
    for path in sorted(_PROFILES_DIR.glob("*.yaml")):
        if path.name == "facets.yaml":
            continue
        raw = _load_yaml(path)
        profile_raw = raw.get("profile")
        profile: dict[str, Any] = profile_raw if isinstance(profile_raw, dict) else {}
        profiles[path.stem] = ProfileInfo(
            key=path.stem,
            name=str(profile.get("name") or _title_from_key(path.stem)),
            description=str(profile.get("description") or "").strip(),
            facets=[str(facet) for facet in profile.get("facets", [])],
        )
    return profiles


def _load_capabilities(path: Path | None) -> CapabilitiesInfo | None:
    raw = _load_yaml(path) if path else {}
    if not raw:
        return None
    server_raw = raw.get("server")
    server: dict[str, Any] = server_raw if isinstance(server_raw, dict) else {}
    return CapabilitiesInfo(
        server_name=str(server.get("name") or "Server under test"),
        active_profile=str(raw.get("active_profile") or ""),
        supported_facets=[str(facet) for facet in raw.get("supported_facets", [])],
        overrides={str(key): str(value) for key, value in (raw.get("cu_overrides") or {}).items()},
    )


def _title_from_key(key: str) -> str:
    acronyms = {"cu": "CU", "id": "ID", "io": "IO", "ijt": "IJT"}
    return " ".join(acronyms.get(token, token.capitalize()) for token in key.split("_"))


def _supported_set(cu_payload: dict[str, Any]) -> set[str] | None:
    supported = cu_payload.get("supported_cus")
    if supported is None:
        return None
    return {str(cu) for cu in supported}


def _server_profile_cu_count(cu_keys: list[str], supported: set[str] | None) -> int | str:
    if supported is None:
        return _NOT_APPLICABLE
    return len([cu_key for cu_key in cu_keys if cu_key in supported])


def _in_server_profile(cu_key: str, supported: set[str] | None) -> str:
    if supported is None:
        return _NOT_APPLICABLE
    return "Yes" if cu_key in supported else "No"


def _server_profile_pct(server_profile_cus: int | str, total: int) -> str:
    if not isinstance(server_profile_cus, int):
        return _NOT_APPLICABLE
    return _pct(server_profile_cus, total)


def _server_supported_cu_keys(cu_keys: list[str], supported: set[str] | None) -> list[str]:
    if supported is None:
        return []
    return [cu_key for cu_key in cu_keys if cu_key in supported]


def _supported_cus_validated_count(cu_keys: list[str], by_cu: dict[str, Any], supported: set[str] | None) -> int | str:
    if supported is None:
        return _NOT_APPLICABLE
    counts = _count_cu_outcomes(_server_supported_cu_keys(cu_keys, supported), by_cu)
    return counts["supported"] + counts["partial"]


def _supported_cus_validated_pct(cu_keys: list[str], by_cu: dict[str, Any], supported: set[str] | None) -> str:
    server_supported_count = _server_profile_cu_count(cu_keys, supported)
    validated_count = _supported_cus_validated_count(cu_keys, by_cu, supported)
    if not isinstance(server_supported_count, int) or not isinstance(validated_count, int):
        return _NOT_APPLICABLE
    return _pct(validated_count, server_supported_count)


def _supported_cus_validated_pct_value(
    cu_keys: list[str], by_cu: dict[str, Any], supported: set[str] | None
) -> float | None:
    server_supported_count = _server_profile_cu_count(cu_keys, supported)
    validated_count = _supported_cus_validated_count(cu_keys, by_cu, supported)
    if not isinstance(server_supported_count, int) or not isinstance(validated_count, int):
        return None
    return _pct_value(validated_count, server_supported_count)


def _run_logs_url() -> str:
    server = os.environ.get("GITHUB_SERVER_URL")
    repo = os.environ.get("GITHUB_REPOSITORY")
    run_id = os.environ.get("GITHUB_RUN_ID")
    if server and repo and run_id:
        return f"{server}/{repo}/actions/runs/{run_id}"
    return _NOT_APPLICABLE


def _package_version(package: str) -> str:
    try:
        return metadata.version(package)
    except metadata.PackageNotFoundError:
        return "not installed"


def _write_baseline(path: Path, baseline: dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(baseline, indent=2, sort_keys=True) + "\n", encoding="utf-8")


def _cus_for_profile(profile: ProfileInfo, facets: dict[str, FacetInfo]) -> list[str]:
    seen: set[str] = set()
    ordered: list[str] = []
    for facet_key in profile.facets:
        for cu_key in facets.get(facet_key, FacetInfo(facet_key, facet_key, "", "", "facet", [])).conformance_units:
            if cu_key not in seen:
                seen.add(cu_key)
                ordered.append(cu_key)
    return ordered


def _cu_to_facets(facets: dict[str, FacetInfo]) -> dict[str, list[str]]:
    mapping: dict[str, list[str]] = {}
    rollups: list[FacetInfo] = []
    for facet in facets.values():
        if facet.kind == "rollup":
            rollups.append(facet)
            continue
        for cu_key in facet.conformance_units:
            mapping.setdefault(cu_key, []).append(facet.display_name)
    for facet in rollups:
        for cu_key in facet.conformance_units:
            if cu_key not in mapping:
                mapping.setdefault(cu_key, []).append(facet.display_name)
    return mapping


def _ordered_cu_keys(by_cu: dict[str, Any], facets: dict[str, FacetInfo]) -> list[str]:
    seen: set[str] = set()
    ordered: list[str] = []
    for facet in facets.values():
        for cu_key in facet.conformance_units:
            if cu_key in by_cu and cu_key not in seen:
                seen.add(cu_key)
                ordered.append(cu_key)
    ordered.extend(sorted(str(cu_key) for cu_key in by_cu if str(cu_key) not in seen))
    return ordered


def _cu_compliance_key(data: dict[str, Any]) -> str:
    explicit = str(data.get("compliance") or "")
    if explicit in _CU_COMPLIANCE_KEYS:
        return explicit

    passed = int(data.get("passed", 0) or 0)
    failed = int(data.get("failed", 0) or 0) + int(data.get("error", 0) or 0)
    not_supported = int(data.get("not_supported", 0) or 0)
    blocked = int(data.get("blocked", 0) or 0)
    accepted_policy = int(data.get("accepted_policy", 0) or 0)
    environment = int(data.get("environment", 0) or 0)
    skipped = int(data.get("skipped", 0) or 0)
    untested = int(data.get("untested", 0) or 0)
    test_count = int(data.get("test_count", 0) or 0)

    if failed:
        return "action_needed"
    if passed and (not_supported or blocked or skipped or untested):
        return "partial"
    if blocked:
        return "blocked"
    if not_supported:
        return "not_supported"
    if passed:
        return "supported"
    if accepted_policy or environment:
        return "blocked"
    if untested or test_count == 0:
        return "untested"
    return "unknown"


def _cu_compliance_label(status: str) -> str:
    return _outcome_label(status)


def _count_cu_outcomes(cu_keys: list[str], by_cu: dict[str, Any]) -> Counter[str]:
    counts: Counter[str] = Counter()
    for cu_key in cu_keys:
        data_raw = by_cu.get(cu_key)
        data: dict[str, Any] = data_raw if isinstance(data_raw, dict) else {}
        counts[_cu_compliance_key(data)] += 1
    return counts


def _reason_bucket(message: str) -> str:
    msg = re.sub(r"\s+", " ", message.strip())
    if not msg:
        return "no reason"
    if msg.startswith("("):
        try:
            longrepr = ast.literal_eval(msg)
        except (SyntaxError, ValueError):
            longrepr = None
        if isinstance(longrepr, tuple) and longrepr and isinstance(longrepr[-1], str):
            msg = re.sub(r"\s+", " ", longrepr[-1].strip())
    msg = re.sub(r"^Skipped:\s*", "", msg, flags=re.IGNORECASE)
    msg = re.split(r"\s+-\s+CU:\s*", msg, maxsplit=1)[0].strip()
    msg = re.split(r"\.\s+To enable:\s*", msg, maxsplit=1)[0].strip()
    msg = re.split(r"\s+Config file:\s*", msg, maxsplit=1)[0].strip()
    not_supported = " NOT SUPPORTED"
    if not_supported in msg:
        end = msg.find(not_supported) + len(not_supported)
        return msg[:end].strip()
    if len(msg) > 180:
        return f"{msg[:177].rstrip()}..."
    return msg


def _skip_diagnostic_category(message: str) -> str:
    """Return the user-facing diagnostic category for one skipped test."""
    normalized = re.sub(r"\s+", " ", str(message or "").strip())
    normalized = re.sub(r"^Skipped:\s*", "", normalized, flags=re.IGNORECASE)
    reason_lower = normalized.lower()
    if reason_lower.startswith("tooling limitation -"):
        return "Test Tooling Limitations"
    if (
        "asyncua addnodes service call unavailable" in reason_lower
        or "asyncua deletenodes service call unavailable" in reason_lower
    ):
        return "Test Tooling Limitations"
    if reason_lower.startswith("companion spec profile note -"):
        return "Companion Spec Profile Notes"
    if reason_lower.startswith("monitoring.") or "lifetimecounters" in reason_lower:
        return "Companion Spec Profile Notes"
    if reason_lower.startswith("simulator regression limit -"):
        return "Simulator Regression Limits"
    if "simulatebulkresults" in reason_lower and (
        "badtoomanyoperations" in reason_lower
        or "concurrent access limit" in reason_lower
        or "stability guard" in reason_lower
    ):
        return "Simulator Regression Limits"
    if "simulatebulkevents" in reason_lower and (
        "badtoomanyoperations" in reason_lower
        or "subscription limit" in reason_lower
        or "stability guard" in reason_lower
    ):
        return "Simulator Regression Limits"
    return "Other Diagnostics"


def _cu_test_index(cu_payload: dict[str, Any]) -> dict[str, list[dict[str, Any]]]:
    indexed: dict[str, list[dict[str, Any]]] = {}
    tests = cu_payload.get("tests")
    if not isinstance(tests, list):
        return indexed
    for test in tests:
        if not isinstance(test, dict):
            continue
        for cu_key in test.get("cus", []):
            indexed.setdefault(str(cu_key), []).append(test)
    return indexed


def _cu_note_summary(cu_key: str, tests_by_cu: dict[str, list[dict[str, Any]]]) -> str:
    tests = tests_by_cu.get(cu_key, [])
    priority = [
        ("failed", "Failed"),
        ("error", "Error"),
        ("blocked", _outcome_label("blocked")),
        ("not_supported", _outcome_label("not_supported")),
        ("accepted_policy", "Accepted Policy"),
        ("environment", "Environment"),
        ("skipped", "Skipped"),
        ("xfailed", "Expected Failure"),
        ("untested", _outcome_label("untested")),
    ]
    notes: list[str] = []
    for outcome, label in priority:
        matching = [test for test in tests if str(test.get("outcome") or "") == outcome]
        if not matching:
            continue
        reason = next((_reason_bucket(str(test.get("reason") or "")) for test in matching if test.get("reason")), "")
        if reason and reason != "no reason":
            notes.append(f"{label}: {reason}")
        else:
            notes.append(f"{label}: {len(matching)} test(s)")
    if not notes:
        return ""
    if len(notes) > 2:
        return "; ".join(notes[:2]) + f"; {len(notes) - 2} more"
    return "; ".join(notes)


def _test_status_label(status: str) -> str:
    """Return the user-facing Excel label for a pytest outcome."""
    return {
        "passed": "Passed",
        "failed": "Failed",
        "error": "Error",
        "skipped": "Skipped",
        "xfailed": "Expected Failure",
        "xpassed": "Unexpected Pass",
    }.get(status, status.replace("_", " ").title())


def _build_report_context(
    cu_payload: dict[str, Any] | None,
    profiles: dict[str, ProfileInfo],
    facets: dict[str, FacetInfo],
    capabilities: CapabilitiesInfo | None,
) -> dict[str, Any] | None:
    if not cu_payload:
        return None
    by_cu = cu_payload.get("by_cu", {}) if isinstance(cu_payload.get("by_cu"), dict) else {}
    supported = _supported_set(cu_payload)
    active_profile = capabilities.active_profile if capabilities else ""
    active = profiles.get(active_profile)
    active_cus = _cus_for_profile(active, facets) if active else []
    active_cus_set = set(active_cus)
    active_counts = _count_cu_outcomes(active_cus, by_cu) if active else Counter()
    server_supported_count = _server_profile_cu_count(active_cus, supported)
    score = _conformance_score(active_counts, server_supported_count, len(active_cus))
    validation_health_value = _supported_cus_validated_pct_value(active_cus, by_cu, supported)
    spec_coverage_value = (
        _pct_value(server_supported_count, len(active_cus)) if isinstance(server_supported_count, int) else None
    )
    ordered_keys = _ordered_cu_keys(by_cu, facets)
    facet_map = _cu_to_facets(facets)
    tests_by_cu = _cu_test_index(cu_payload)

    findings: list[dict[str, Any]] = []
    cu_outcomes: dict[str, str] = {}
    for cu_key in ordered_keys:
        data_raw = by_cu.get(cu_key)
        data = data_raw if isinstance(data_raw, dict) else {}
        outcome = _cu_compliance_key(data)
        cu_outcomes[cu_key] = outcome
        if outcome not in _FINDING_OUTCOMES:
            continue
        status, status_icon = _status_for(cu_key, outcome, active_cus_set)
        findings.append(
            {
                "cu_key": cu_key,
                "cu": cu_display_name(cu_key),
                "facets": ", ".join(facet_map.get(cu_key, [])),
                "outcome": outcome,
                "result": _cu_compliance_label(outcome),
                "reason": _cu_note_summary(cu_key, tests_by_cu),
                "status": status,
                "status_icon": status_icon,
            }
        )
    findings.sort(key=lambda item: (_STATUS_ORDER.get(str(item["status"]), 99), str(item["cu_key"])))
    findings_count = Counter(_status_count_key(str(item["status"])) for item in findings)
    is_healthy = _is_healthy(
        context_present=True,
        server_supported_count=server_supported_count,
        active_cus_len=len(active_cus),
        failed_count=int(findings_count.get("action_needed", 0) or 0),
        blocked_count=int(findings_count.get("blocked", 0) or 0),
    )
    return {
        "by_cu": by_cu,
        "supported": supported,
        "active_profile": active_profile,
        "active_label": active.name if active else "No active profile found",
        "active_cus": active_cus,
        "active_counts": active_counts,
        "server_supported_count": server_supported_count,
        "score": score,
        "validation_health_value": validation_health_value,
        "spec_coverage_value": spec_coverage_value,
        "findings": findings,
        "findings_count": findings_count,
        "is_healthy": is_healthy,
        "cu_outcomes": cu_outcomes,
    }


def _baseline_payload(context: dict[str, Any], run_ts_iso: str) -> dict[str, Any]:
    return {
        "run_ts": run_ts_iso,
        "git_sha": _short_git_sha(_PROJECT_ROOT),
        "score": context["score"],
        "validation_health_pct": context["validation_health_value"],
        "spec_coverage_pct": context["spec_coverage_value"],
        "findings_count": dict(context["findings_count"]),
        "cu_outcomes": context["cu_outcomes"],
    }


def _support_rows(context: dict[str, Any], facets: dict[str, FacetInfo], limit: int = 12) -> list[tuple[str, str]]:
    by_cu: dict[str, Any] = context["by_cu"]
    supported: set[str] | None = context["supported"]
    rows: list[tuple[int, str, str]] = []
    for facet in facets.values():
        if facet.kind == "rollup":
            continue
        counts = _count_cu_outcomes(facet.conformance_units, by_cu)
        server_supported_count = _server_profile_cu_count(facet.conformance_units, supported)
        if counts["action_needed"] or counts["blocked"] or server_supported_count == 0:
            icon, rank, label = _CAPABILITY_SUPPORT_ICONS["not_supported"], 0, "Not Supported"
        elif counts["partial"] or counts["not_supported"] or server_supported_count != len(facet.conformance_units):
            icon, rank, label = _CAPABILITY_SUPPORT_ICONS["partial"], 1, "Partially Supported"
        else:
            icon, rank, label = _CAPABILITY_SUPPORT_ICONS["supported"], 2, "Supported"
        name = facet.display_name.removeprefix("IJT ").removesuffix(" Server Facet")
        rows.append((rank, name, f"{icon} {label}. {facet.description}"))
    rows.sort(key=lambda item: (item[0], item[1]))
    return [(name, status) for _rank, name, status in rows[:limit]]


def _compliance_status(counts: Counter[str], total: int) -> str:
    if counts["action_needed"]:
        return "Failed"
    if counts["blocked"]:
        return "Blocked"
    if counts["partial"] or counts["not_supported"]:
        return "Supported With Notes"
    if total and counts["supported"] == total:
        return "Supported"
    if counts["supported"]:
        return "Supported With Notes"
    return "No Compliance Result"


def _status_fill(status: str) -> PatternFill:
    colour = {
        "Supported": _LIGHT_GREEN,
        "Supported With Notes": _LIGHT_GREEN_NOTE,
        "Blocked": _LIGHT_ORANGE,
        "Failed": _LIGHT_RED,
        "No Compliance Result": _GRAY,
    }.get(status, _WHITE)
    return _fill(colour)


def _review_status_fill(status: str) -> PatternFill:
    return _fill(_status_color_excel(status))


def _percentage_fill(value: str | float | None) -> PatternFill:
    if value is None or value == _NOT_APPLICABLE:
        return _fill(_GRAY)
    if isinstance(value, str):
        try:
            pct = float(value.removesuffix("%"))
        except ValueError:
            return _fill(_WHITE)
    else:
        pct = value
    if pct >= 99.9:
        return _fill(_LIGHT_GREEN)
    if pct >= 90.0:
        return _fill(_LIGHT_GREEN_NOTE)
    if pct >= 75.0:
        return _fill(_LIGHT_YELLOW)
    if pct >= 50.0:
        return _fill(_LIGHT_ORANGE)
    return _fill(_LIGHT_RED)


def _pct(numerator: int, denominator: int) -> str:
    if denominator <= 0:
        return _NOT_APPLICABLE
    return f"{(numerator * 100 / denominator):.1f}%"


# ── Excel helpers ─────────────────────────────────────────────────────────────


def _fill(hex_colour: str) -> PatternFill:
    return PatternFill(fill_type="solid", fgColor=hex_colour)


def _header_font() -> Font:
    return Font(bold=True, color="FF000000")


def _apply_print_setup(ws, *, repeat_header_row: int | None = 1) -> None:
    """Apply uniform landscape / A4 / fit-to-width print settings to *ws*.

    ``repeat_header_row`` is the 1-based spreadsheet row to repeat as a print
    title on every page. Pass ``None`` for layout sheets without a single
    repeating table header row (e.g. Conformance Overview).
    """
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0  # unbounded height; only width fits to 1 page
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.print_options.horizontalCentered = True
    ws.page_margins.top = 0.5
    ws.page_margins.bottom = 0.5
    ws.page_margins.left = 0.3
    ws.page_margins.right = 0.3
    ws.page_margins.header = 0.3
    ws.page_margins.footer = 0.3
    if repeat_header_row is not None and ws.max_row >= repeat_header_row:
        ws.print_title_rows = f"{repeat_header_row}:{repeat_header_row}"
    ws.oddHeader.center.text = "IJT Conformance Test Report"
    ws.oddHeader.center.size = 12
    ws.oddHeader.center.color = "24292E"
    ws.oddFooter.center.text = "Page &P of &N"
    ws.oddFooter.center.size = 10


def _apply_autofilter(ws, *, start_row: int = 1) -> None:
    """Set auto-filter from ``start_row`` (the table header) through the last data row.

    ``start_row`` is the 1-based spreadsheet row holding the real table header.
    Skipped when the sheet has no data rows below the header.
    """
    if ws.max_row <= start_row:
        return
    last_col_letter = openpyxl.utils.get_column_letter(ws.max_column)
    last_row = ws.max_row
    ws.auto_filter.ref = f"A{start_row}:{last_col_letter}{last_row}"


def _apply_header(ws, row: int, headers: list[str], col_widths: list[int]) -> None:
    for col, (hdr, width) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=row, column=col, value=hdr)
        cell.font = _header_font()
        cell.fill = _fill(_BLUE)
        cell.alignment = Alignment(wrap_text=True, vertical="center")
        ws.column_dimensions[get_column_letter(col)].width = width


def _write_row(ws, row: int, values: list, status: str, alternate: bool) -> None:
    bg = _fill(_GRAY) if alternate else _fill(_WHITE)
    for col, val in enumerate(values, start=1):
        cell = ws.cell(row=row, column=col, value=val)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        if col == 4:  # Status column — colour by status
            cell.fill = _fill(_STATUS_COLOUR.get(status, _WHITE))
            cell.font = Font(bold=True)
        else:
            cell.fill = bg


# ── Sheet builders ────────────────────────────────────────────────────────────


def _write_failure_banner(ws, row: int) -> None:
    ws.cell(row=row, column=1, value="Run result").font = Font(bold=True)
    cell = ws.cell(
        row=row,
        column=2,
        value=(
            "Failed - this workbook was generated for diagnostics after a failed test run. "
            "Coverage may be partial; review the Failures sheet and runner output first."
        ),
    )
    cell.fill = _fill(_LIGHT_RED)
    cell.font = Font(bold=True)
    cell.alignment = Alignment(wrap_text=True, vertical="top")


def _build_cover(
    wb: openpyxl.Workbook,
    cases: list[TestCase],
    run_ts: str,
    run_result: str,
    context: dict[str, Any] | None,
    facets: dict[str, FacetInfo],
) -> None:
    ws = wb.create_sheet("Conformance Overview")
    ws.sheet_view.showGridLines = False
    for col in range(1, 8):
        ws.column_dimensions[get_column_letter(col)].width = 22

    statuses = [case.status for case in cases]
    failed = statuses.count("failed") + statuses.count("error")
    status = "Passed" if failed == 0 and run_result != "failed" else "Failed"
    _status_emoji = "🟢" if status == "Passed" else "🔴"
    ws.merge_cells("A1:G1")
    title = ws["A1"]
    if context:
        supported = context["server_supported_count"]
        total_active = len(context["active_cus"])
        validated = _supported_cus_validated_count(context["active_cus"], context["by_cu"], context["supported"])
        findings_count: Counter[str] = context["findings_count"]
        action_total = int(findings_count.get("action_needed", 0) or 0) + int(findings_count.get("blocked", 0) or 0)
        action_label = "action item" if action_total == 1 else "action items"
        title.value = (
            f"{_status_emoji} {status} · {action_total} {action_label} · "
            f"Validation {_fmt_pct(context['validation_health_value'])} ({validated}/{supported}) · "
            f"Server support {_fmt_pct(context['spec_coverage_value'])} ({supported}/{total_active})"
        )
    else:
        title.value = f"{_status_emoji} {status} · conformance profile unavailable"
    title.font = Font(bold=True, size=24)
    title.fill = _fill(_LIGHT_GREEN if status == "Passed" else _LIGHT_RED)
    title.alignment = Alignment(horizontal="center")

    if run_result == "failed":
        _write_failure_banner(ws, 3)

    row = 5
    _apply_header(ws, row, ["Metric", "Value", "Context"], [28, 24, 70])
    if context:
        supported = context["server_supported_count"]
        total_active = len(context["active_cus"])
        validated = _supported_cus_validated_count(context["active_cus"], context["by_cu"], context["supported"])
        findings_count = context["findings_count"]
        metrics = [
            (
                "Server Support Coverage",
                _fmt_pct(context["spec_coverage_value"]),
                f"{supported} / {total_active} CUs server-supported",
            ),
            (
                "Validation Health",
                _fmt_pct(context["validation_health_value"]),
                f"{validated} / {supported} server-supported CUs validated",
            ),
        ]
        if not bool(context.get("is_healthy")):
            metrics.append(
                (
                    "Conformance Action Items",
                    _format_status_counts(_ACTION_ITEM_LABEL_ORDER, findings_count),
                    _action_items_context(findings_count),
                )
            )
        metrics.append(
            (
                "Server Scope Notes",
                _format_status_counts(_CAPABILITY_NOTE_LABEL_ORDER, findings_count),
                _informational_notes_context(findings_count),
            )
        )
        for offset, (metric, value, note) in enumerate(metrics, start=1):
            ws.cell(row=row + offset, column=1, value=metric).font = Font(bold=True)
            value_cell = ws.cell(row=row + offset, column=2, value=value)
            value_cell.font = Font(bold=True)
            if metric in {"Server Support Coverage", "Validation Health"}:
                value_cell.fill = _percentage_fill(
                    context["spec_coverage_value" if metric == "Server Support Coverage" else "validation_health_value"]
                )
            ws.cell(row=row + offset, column=3, value=note).alignment = Alignment(wrap_text=True, vertical="top")

    row = 10
    ws.cell(row=row, column=1, value="IJT Facet Support").font = Font(bold=True, size=14)
    _apply_header(ws, row + 1, ["IJT Facet", "Status"], [34, 90])
    if context:
        for offset, (area, status_text) in enumerate(_support_rows(context, facets), start=2):
            ws.cell(row=row + offset, column=1, value=area).font = Font(bold=True)
            ws.cell(row=row + offset, column=2, value=status_text).alignment = Alignment(wrap_text=True, vertical="top")

    row = row + 16
    ws.cell(row=row, column=1, value="Test Client Environment").font = Font(bold=True, size=14)
    env_rows = [
        ("Generated", run_ts),
        ("Commit", _short_git_sha(_PROJECT_ROOT)),
        ("Python", platform.python_version()),
        ("asyncua", _package_version("asyncua")),
        ("Host OS", platform.platform()),
        ("Run logs", _run_logs_url()),
    ]
    _apply_header(ws, row + 1, ["Item", "Value"], [28, 80])
    for offset, (item, value) in enumerate(env_rows, start=2):
        ws.cell(row=row + offset, column=1, value=item).font = Font(bold=True)
        ws.cell(row=row + offset, column=2, value=value).alignment = Alignment(wrap_text=True, vertical="top")

    _apply_print_setup(ws, repeat_header_row=None)  # layout sheet — no single repeating header row
    ws.freeze_panes = "A3"


def _build_summary(wb: openpyxl.Workbook, cases: list[TestCase], run_ts: str, run_result: str) -> None:
    ws = wb.create_sheet("Test Outcome Counts")
    ws.sheet_view.showGridLines = False

    # Title
    ws["A1"] = "IJT Test Client — Conformance Test Report"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = f"Generated: {run_ts}"
    ws["A2"].font = Font(italic=True, size=10)
    if run_result == "failed":
        _write_failure_banner(ws, 3)

    # Overall counts
    statuses = [c.status for c in cases]
    passed = statuses.count("passed")
    failed = statuses.count("failed") + statuses.count("error")
    skipped = statuses.count("skipped")
    xfailed = statuses.count("xfailed") + statuses.count("xpassed")
    total = len(cases)

    ws["A4"] = "Overall"
    ws["A4"].font = Font(bold=True)
    summary_data = [
        ("Total", total, _BLUE),
        ("Passed", passed, _GREEN),
        ("Failed", failed, _RED),
        ("Skipped", skipped, _YELLOW),
        ("Xfailed", xfailed, _ORANGE),
    ]
    for i, (label, count, colour) in enumerate(summary_data, start=5):
        ws.cell(row=i, column=1, value=label).font = Font(bold=True)
        cell = ws.cell(row=i, column=2, value=count)
        cell.fill = _fill(colour)
        cell.font = Font(bold=True)

    # Per-area breakdown
    ws["A11"] = "By Test Area"
    ws["A11"].font = Font(bold=True)
    _apply_header(ws, 12, ["Area", "Total", "Passed", "Failed", "Skipped", "Xfailed"], [20, 8, 8, 8, 10, 10])

    areas: dict[str, list[TestCase]] = {}
    for c in cases:
        areas.setdefault(c.area, []).append(c)

    for row_idx, (area, area_cases) in enumerate(sorted(areas.items()), start=13):
        sc = [c.status for c in area_cases]
        values = [
            area,
            len(area_cases),
            sc.count("passed"),
            sc.count("failed") + sc.count("error"),
            sc.count("skipped"),
            sc.count("xfailed") + sc.count("xpassed"),
        ]
        for col, val in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            if col == 4 and isinstance(val, int) and val > 0:
                cell.fill = _fill(_RED)
                cell.font = Font(bold=True)
            elif col == 3:
                cell.fill = _fill(_GREEN if val == len(area_cases) else _WHITE)

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 8

    _apply_print_setup(ws, repeat_header_row=12)  # real table header (`By Test Area`) lives at row 12
    _apply_autofilter(ws, start_row=12)


def _build_all_tests(wb: openpyxl.Workbook, cases: list[TestCase]) -> None:
    ws = wb.create_sheet("All Test Cases")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A2"

    headers = ["Area", "File", "Test Name", "Status", "Duration (s)", "Reason / Message"]
    col_widths = [15, 35, 55, 10, 12, 70]
    _apply_header(ws, 1, headers, col_widths)

    for row_idx, (i, c) in enumerate(enumerate(cases), start=2):
        _write_row(
            ws,
            row_idx,
            [
                c.area,
                c.file,
                c.short_name,
                _test_status_label(c.status),
                round(c.duration, 3),
                c.message,
            ],
            c.status,
            i % 2 == 0,
        )

    ws.row_dimensions[1].height = 20

    _apply_print_setup(ws)
    _apply_autofilter(ws)


def _build_filtered(
    wb: openpyxl.Workbook, cases: list[TestCase], sheet_name: str, statuses: list[str], colour: str
) -> None:
    filtered = [c for c in cases if c.status in statuses]
    include_skip_category = statuses == ["skipped"] and sheet_name == "Skipped Test Cases"
    ws = wb.create_sheet(f"{sheet_name} ({len(filtered)})")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A2"

    if not filtered:
        ws["A1"] = f"No {sheet_name.lower()} tests."
        _apply_print_setup(ws)
        _apply_autofilter(ws)  # skipped automatically — max_row == 1
        return

    headers = ["Area", "File", "Test Name", "Duration (s)", "Reason / Message"]
    col_widths = [15, 35, 55, 12, 80]
    if include_skip_category:
        headers = ["Area", "File", "Test Name", "Duration (s)", "Category", "Reason / Message"]
        col_widths = [15, 35, 55, 12, 30, 80]
    _apply_header(ws, 1, headers, col_widths)

    for row_idx, c in enumerate(filtered, start=2):
        values = [c.area, c.file, c.short_name, round(c.duration, 3), c.message]
        if include_skip_category:
            values = [
                c.area,
                c.file,
                c.short_name,
                round(c.duration, 3),
                _skip_diagnostic_category(c.message),
                c.message,
            ]
        for col, val in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.fill = _fill(colour)
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    _apply_print_setup(ws)
    _apply_autofilter(ws)


def _write_metric_block(ws, start_row: int, rows: list[tuple[str, object]]) -> int:
    _apply_header(ws, start_row, ["Metric", "Value"], [34, 80])
    for idx, (label, value) in enumerate(rows, start=start_row + 1):
        ws.cell(row=idx, column=1, value=label).font = Font(bold=True)
        ws.cell(row=idx, column=2, value=value).alignment = Alignment(wrap_text=True, vertical="top")
    return start_row + len(rows) + 2


def _build_profile_coverage(
    wb: openpyxl.Workbook,
    cu_payload: dict[str, Any],
    profiles: dict[str, ProfileInfo],
    facets: dict[str, FacetInfo],
    capabilities: CapabilitiesInfo | None,
    run_result: str,
) -> None:
    ws = wb.create_sheet("Profile Coverage Comparison")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A10"

    by_cu = cu_payload.get("by_cu", {}) if isinstance(cu_payload.get("by_cu"), dict) else {}
    supported = _supported_set(cu_payload)
    summary = cu_payload.get("summary", {}) if isinstance(cu_payload.get("summary"), dict) else {}
    all_cu_keys = _ordered_cu_keys(by_cu, facets)

    ws["A1"] = "IJT Profile, Facet, and Conformance Unit Coverage"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = (
        "This sheet maps the current test run to IJT coverage views, facets, and conformance units. "
        "Primary reader view: see IJT Facet Breakdown. This sheet retains the active profile and three reference views. "
        "Start with the Server Capability Profile row; Reference IJT Facet and Reference Full CU Set rows are comparison views only, not extra pass/fail requirements. "
        "Server Supported CUs comes from the server capability file. Outcome and validated counts come from this test run. "
        "Supported CUs Validated % is the main health signal and is color-coded; Server Support % is informational and is not color-coded. "
        "Skip diagnostics may overlap with Conformance Action Items."
    )
    ws["A2"].alignment = Alignment(wrap_text=True)
    if run_result == "failed":
        _write_failure_banner(ws, 3)

    active = capabilities.active_profile if capabilities else ""
    server = capabilities.server_name if capabilities else "Server under test"
    next_row = _write_metric_block(
        ws,
        4,
        [
            ("Server", server),
            ("Server capability profile", profiles.get(active, ProfileInfo(active, active or "Unknown", "", [])).name),
            ("Official IJT CUs", summary.get("official_cu_count", len(by_cu))),
            ("Server supported CUs", len(supported) if supported is not None else _NOT_APPLICABLE),
            ("Workbook test cases", summary.get("workbook_case_count", _NOT_APPLICABLE)),
        ],
    )

    headers = [
        "View",
        "Role",
        "Facet Count",
        "CUs in View",
        "Server Supported CUs",
        "Validated Supported",
        "Validated with Notes",
        _outcome_label("not_supported"),
        _outcome_label("blocked"),
        _outcome_label("action_needed"),
        _outcome_label("untested"),
        "Server Support %",
        "Supported CUs Validated %",
        "Outcome",
        "Description",
    ]
    widths = [34, 28, 12, 10, 14, 12, 12, 14, 9, 13, 9, 12, 14, 16, 70]
    _apply_header(ws, next_row, headers, widths)

    view_rows: list[tuple[str, str, int, list[str], str]] = []
    if active in profiles:
        profile = profiles[active]
        view_rows.append(
            (
                profile.name,
                "Server Capability Profile",
                len(profile.facets),
                _cus_for_profile(profile, facets),
                profile.description,
            )
        )

    for facet_key in (
        "basic_joining_system_server_facet",
        "general_joining_system_server_facet",
        "joining_system_selectable_features_server_facet",
    ):
        facet = facets.get(facet_key)
        if facet is None:
            continue
        view_rows.append(
            (
                facet.display_name,
                "Reference IJT Facet",
                1,
                facet.conformance_units,
                facet.description,
            )
        )

    if active != "full_conformance":
        view_rows.append(
            (
                "Full IJT Base CU Set",
                "Reference Full CU Set",
                len(facets),
                all_cu_keys,
                "Flat reference view of all unique IJT Base conformance units.",
            )
        )

    row = next_row + 1
    for view_name, role, facet_count, cu_keys, description in view_rows:
        counts = _count_cu_outcomes(cu_keys, by_cu)
        server_profile_cus = _server_profile_cu_count(cu_keys, supported)
        supported_validated_pct_value = _supported_cus_validated_pct_value(cu_keys, by_cu, supported)
        compliance = _compliance_status(counts, len(cu_keys))
        values = [
            view_name,
            role,
            facet_count,
            len(cu_keys),
            server_profile_cus,
            counts["supported"],
            counts["partial"],
            counts["not_supported"],
            counts["blocked"],
            counts["action_needed"],
            counts["untested"],
            _server_profile_pct(server_profile_cus, len(cu_keys)),
            _supported_cus_validated_pct(cu_keys, by_cu, supported),
            compliance,
            description,
        ]
        for col, value in enumerate(values, start=1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            if col == 2 and value == "Server Capability Profile":
                cell.fill = _fill(_LIGHT_BLUE)
                cell.font = Font(bold=True)
            if col == 13:
                cell.fill = _percentage_fill(supported_validated_pct_value)
                cell.font = Font(bold=True)
            if col == 14:
                cell.fill = _status_fill(str(value))
                cell.font = Font(bold=True)
        row += 1

    _apply_print_setup(ws, repeat_header_row=next_row)  # real table header lives below the metric block
    _apply_autofilter(ws, start_row=next_row)


def _build_facet_coverage(
    wb: openpyxl.Workbook,
    cu_payload: dict[str, Any],
    facets: dict[str, FacetInfo],
    capabilities: CapabilitiesInfo | None,
) -> None:
    ws = wb.create_sheet("IJT Facet Breakdown")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A2"

    by_cu = cu_payload.get("by_cu", {}) if isinstance(cu_payload.get("by_cu"), dict) else {}
    supported = _supported_set(cu_payload)
    active_extra_facets = set(capabilities.supported_facets if capabilities else [])

    headers = [
        "Facet",
        "Facet Key",
        "Facet Type",
        "CUs in Facet",
        "Server Supported CUs",
        "Validated Supported",
        "Validated with Notes",
        _outcome_label("not_supported"),
        _outcome_label("blocked"),
        _outcome_label("action_needed"),
        _outcome_label("untested"),
        "Server Support %",
        "Supported CUs Validated %",
        "Outcome",
        "Description",
    ]
    widths = [34, 34, 12, 10, 14, 12, 12, 14, 9, 13, 9, 12, 14, 16, 70]
    _apply_header(ws, 1, headers, widths)

    for row, facet in enumerate(facets.values(), start=2):
        cu_keys = facet.conformance_units
        counts = _count_cu_outcomes(cu_keys, by_cu)
        server_profile_cus = _server_profile_cu_count(cu_keys, supported)
        supported_validated_pct_value = _supported_cus_validated_pct_value(cu_keys, by_cu, supported)
        compliance = _compliance_status(counts, len(cu_keys))
        facet_type = "Facet Group" if facet.kind == "rollup" else "Facet"
        if facet.key in active_extra_facets:
            facet_type = f"Additional {facet_type}"
        values = [
            facet.display_name,
            facet.key,
            facet_type,
            len(cu_keys),
            server_profile_cus,
            counts["supported"],
            counts["partial"],
            counts["not_supported"],
            counts["blocked"],
            counts["action_needed"],
            counts["untested"],
            _server_profile_pct(server_profile_cus, len(cu_keys)),
            _supported_cus_validated_pct(cu_keys, by_cu, supported),
            compliance,
            facet.description,
        ]
        for col, value in enumerate(values, start=1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            if col == 13:
                cell.fill = _percentage_fill(supported_validated_pct_value)
                cell.font = Font(bold=True)
            if col == 14:
                cell.fill = _status_fill(str(value))
                cell.font = Font(bold=True)

    _apply_print_setup(ws)
    _apply_autofilter(ws)


def _build_cu_coverage(
    wb: openpyxl.Workbook,
    cu_payload: dict[str, Any],
    facets: dict[str, FacetInfo],
    capabilities: CapabilitiesInfo | None,
    context: dict[str, Any] | None,
) -> None:
    ws = wb.create_sheet("Conformance Unit Details")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A2"

    by_cu = cu_payload.get("by_cu", {}) if isinstance(cu_payload.get("by_cu"), dict) else {}
    supported = _supported_set(cu_payload)
    facet_map = _cu_to_facets(facets)
    overrides = capabilities.overrides if capabilities else {}
    ordered_keys = _ordered_cu_keys(by_cu, facets)
    tests_by_cu = _cu_test_index(cu_payload)
    active_cus = set(context["active_cus"]) if context else set()

    headers = [
        "Review Status",
        "CU",
        "CU Key",
        "Facet(s)",
        "Server Supported",
        "Outcome",
        "Primary Reason",
        "Tests",
        "Passed",
        _outcome_label("not_supported"),
        _outcome_label("blocked"),
        "Failures",
        "Workbook Cases",
        "Positive",
        "Negative",
        "Override",
        "Notes",
        "Example Test",
    ]
    widths = [18, 34, 34, 44, 18, 16, 60, 8, 8, 14, 9, 12, 14, 9, 9, 14, 80, 80]
    _apply_header(ws, 1, headers, widths)

    for row, cu_key in enumerate(ordered_keys, start=2):
        data_raw = by_cu.get(cu_key)
        data = data_raw if isinstance(data_raw, dict) else {}
        compliance = _cu_compliance_key(data)
        failed = int(data.get("failed", 0) or 0) + int(data.get("error", 0) or 0)
        tests = data.get("tests") if isinstance(data.get("tests"), list) else []
        support = _in_server_profile(cu_key, supported)
        status, status_icon = _status_for(cu_key, compliance, active_cus)
        values: list[Any] = [
            f"{status_icon} {status}",
            cu_display_name(cu_key),
            cu_key,
            ", ".join(facet_map.get(cu_key, [])),
            support,
            _cu_compliance_label(compliance),
            _cu_note_summary(cu_key, tests_by_cu),
            int(data.get("test_count", 0) or 0),
            int(data.get("passed", 0) or 0),
            int(data.get("not_supported", 0) or 0),
            int(data.get("blocked", 0) or 0),
            failed,
            int(data.get("workbook_case_count", 0) or 0),
            int(data.get("workbook_positive_case_count", 0) or 0),
            int(data.get("workbook_negative_case_count", 0) or 0),
            overrides.get(cu_key, ""),
            _cu_note_summary(cu_key, tests_by_cu),
            str(tests[0]) if tests else "",
        ]
        for col, value in enumerate(values, start=1):
            cell = ws.cell(row=row, column=col, value=value)
            header = headers[col - 1]
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            if header == "Review Status" and status:
                cell.font = Font(bold=True)
                cell.fill = _review_status_fill(status)
            if header == "Server Supported" and value == "No":
                cell.fill = _fill(_GRAY)
            if header == "Outcome":
                cell.fill = _fill(_CU_STATUS_COLOUR.get(compliance, _WHITE))
                cell.font = Font(bold=True)

    _apply_print_setup(ws)
    _apply_autofilter(ws)


# ── Main ──────────────────────────────────────────────────────────────────────


def _parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    p.add_argument("--xml", default="test-results/pytest.xml", help="JUnit XML input file")
    p.add_argument("--out", default="test-results/report.xlsx", help="Excel output file")
    p.add_argument(
        "--cu-json",
        default=str(_DEFAULT_CU_JSON),
        help="Optional CU compliance JSON input for profile/facet/CU sheets",
    )
    p.add_argument(
        "--capabilities",
        default=None,
        help="Optional server capabilities YAML used to label the active profile",
    )
    p.add_argument(
        "--run-result",
        choices=["passed", "failed", "unknown"],
        default=os.environ.get("IJT_RUN_RESULT", "unknown"),
        help="Overall runner result; failed adds a diagnostic warning banner to the workbook",
    )
    p.add_argument("--baseline", default=str(_DEFAULT_BASELINE_JSON), help="Report baseline JSON path")
    p.add_argument(
        "--write-baseline",
        action="store_true",
        help="Write report-baseline.json after rendering; CI summary normally owns this in Integration",
    )
    return p.parse_args()


def main() -> int:
    args = _parse_args()
    xml_path = Path(args.xml)
    out_path = Path(args.out)

    if not xml_path.exists():
        print(f"ERROR: JUnit XML not found: {xml_path}", file=sys.stderr)
        print("  Run tests first: python -m pytest ... --junitxml=test-results/pytest.xml", file=sys.stderr)
        return 1

    out_path.parent.mkdir(parents=True, exist_ok=True)

    print(f"Reading: {xml_path}")
    cases = parse_junit_xml(xml_path)
    print(f"  {len(cases)} test cases found")
    cu_payload = _load_json(Path(args.cu_json))
    capabilities_arg = args.capabilities or os.environ.get("OPCUA_CAPABILITIES_FILE")
    capabilities_path = Path(capabilities_arg) if capabilities_arg else _PROJECT_ROOT / "server_capabilities.yaml"
    capabilities = _load_capabilities(capabilities_path)
    facets = _load_facets()
    profiles = _load_profiles()
    baseline_path = Path(args.baseline)
    context = _build_report_context(cu_payload, profiles, facets, capabilities)

    run_ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    run_ts_iso = datetime.now(timezone.utc).isoformat().replace("+00:00", "Z")

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default empty sheet

    _build_cover(wb, cases, run_ts, args.run_result, context, facets)
    _build_summary(wb, cases, run_ts, args.run_result)
    if cu_payload and facets:
        _build_facet_coverage(wb, cu_payload, facets, capabilities)
        _build_cu_coverage(wb, cu_payload, facets, capabilities, context)
        _build_profile_coverage(wb, cu_payload, profiles, facets, capabilities, args.run_result)
    _build_all_tests(wb, cases)
    _build_filtered(wb, cases, "Test Failures", ["failed", "error"], _RED)
    _build_filtered(wb, cases, "Skipped Test Cases", ["skipped"], _YELLOW)
    _build_filtered(wb, cases, "Expected Failures", ["xfailed", "xpassed"], _ORANGE)

    wb.save(out_path)
    if args.write_baseline and context is not None:
        _write_baseline(baseline_path, _baseline_payload(context, run_ts_iso))

    # Print summary to console
    statuses = [c.status for c in cases]
    passed = statuses.count("passed")
    failed = statuses.count("failed") + statuses.count("error")
    skipped = statuses.count("skipped")
    xfailed = statuses.count("xfailed") + statuses.count("xpassed")

    print(f"\nReport written: {out_path}")
    print(f"  Passed:   {passed}")
    print(f"  Failed:   {failed}")
    print(f"  Skipped:  {skipped}")
    print(f"  Xfailed:  {xfailed}")
    print(f"  Total:    {len(cases)}")
    if cu_payload and facets:
        print("  CU sheets: IJT Facet Breakdown, Conformance Unit Details, Profile Coverage Comparison")
    if failed > 0:
        print(f"\n  *** {failed} FAILURE(S) — see 'Test Failures' sheet ***")
    return 0


if __name__ == "__main__":
    sys.exit(main())
