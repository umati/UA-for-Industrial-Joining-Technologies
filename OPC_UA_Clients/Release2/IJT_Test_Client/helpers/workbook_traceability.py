"""
Workbook row traceability for IJT conformance reporting.

The Test Cases workbook is the authored compliance surface. Runtime tests use
stable CU keys, so this helper maps each CU worksheet, test-case header row, and
TC number back to those keys without putting spec sequence numbers in test code.
"""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Any, Iterable, Mapping

from openpyxl import load_workbook

from helpers.cu_registry import CU

PROJECT_ROOT = Path(__file__).resolve().parents[1]
DEFAULT_WORKBOOK_PATH = (
    PROJECT_ROOT / "docs" / "TestCases_OPC 40450-1 UA CS for Joining Systems - Part 1 - Base 1.01.0.xlsm"
)
INFRA_SHEET_COUNT = 6
HEADER_COLUMNS = 11
EXPECTED_WORKBOOK_CASE_COUNT = 1122


@dataclass(frozen=True)
class WorkbookCase:
    """One Step=0 test-case header row from a CU worksheet."""

    cu_key: str
    sheet: str
    row: int
    tc_no: int
    title: str
    case_type: str
    ctt: str
    reviewed: str
    spec_link: str

    @property
    def ref(self) -> str:
        return workbook_ref_id(self.sheet, self.row)

    def to_json(self, linked_test_nodeids: Iterable[str], exact_test_nodeids: Iterable[str]) -> dict[str, Any]:
        linked = sorted(set(linked_test_nodeids))
        exact = sorted(set(exact_test_nodeids))
        return {
            "ref": self.ref,
            "sheet": self.sheet,
            "row": self.row,
            "tc_no": self.tc_no,
            "title": self.title,
            "case_type": self.case_type,
            "ctt": self.ctt,
            "reviewed": self.reviewed,
            "spec_link": self.spec_link,
            "linked_test_count": len(linked),
            "linked_tests": linked,
            "exact_test_count": len(exact),
            "exact_tests": exact,
            "mapping_precision": "exact" if exact else "cu",
        }


def ordered_cu_keys() -> tuple[str, ...]:
    """Return CU keys in workbook/spec order using the CU registry declaration order."""
    return tuple(value for name, value in CU.__dict__.items() if name.isupper() and isinstance(value, str))


def workbook_ref_id(sheet: str, row: int) -> str:
    """Return the stable workbook row reference used in JSON artifacts."""
    return f"{sheet}!R{row}"


def _is_step_zero(value: object) -> bool:
    try:
        return int(value) == 0
    except (TypeError, ValueError):
        return False


def _to_int(value: object) -> int | None:
    try:
        return int(value)
    except (TypeError, ValueError):
        return None


def _cell_text(value: object) -> str:
    return "" if value is None else str(value).strip()


def _case_rows_for_sheet(ws, cu_key: str) -> tuple[WorkbookCase, ...]:
    cases: list[WorkbookCase] = []
    in_error_section = False
    for row in ws.iter_rows(min_row=1, max_col=HEADER_COLUMNS):
        values = [cell.value for cell in row]
        first = _cell_text(values[0])
        if first.lower() == "error test cases":
            in_error_section = True
            continue

        no = _to_int(values[1])
        if no is None or not _is_step_zero(values[2]):
            continue

        cases.append(
            WorkbookCase(
                cu_key=cu_key,
                sheet=ws.title,
                row=row[0].row,
                tc_no=no,
                title=_cell_text(values[3]),
                case_type="negative" if in_error_section else "positive",
                ctt=first,
                reviewed=_cell_text(values[8]),
                spec_link=_cell_text(values[9]),
            )
        )
    return tuple(cases)


def load_workbook_cases(workbook_path: Path | None = None) -> dict[str, tuple[WorkbookCase, ...]]:
    """Load all workbook TC header rows grouped by CU key."""
    source = workbook_path or DEFAULT_WORKBOOK_PATH
    cu_keys = ordered_cu_keys()
    workbook = load_workbook(source, read_only=True, keep_vba=True, data_only=False)
    try:
        cu_sheets = workbook.worksheets[INFRA_SHEET_COUNT:]
        if len(cu_sheets) != len(cu_keys):
            raise ValueError(f"Workbook has {len(cu_sheets)} CU sheets but registry has {len(cu_keys)} CU keys")
        return {cu_key: _case_rows_for_sheet(ws, cu_key) for cu_key, ws in zip(cu_keys, cu_sheets, strict=True)}
    finally:
        workbook.close()


def workbook_traceability_report(
    *,
    workbook_path: Path | None = None,
    tests_by_cu: Mapping[str, Iterable[Mapping[str, Any]]] | None = None,
    exact_refs_by_row: Mapping[str, Iterable[str]] | None = None,
) -> dict[str, Any]:
    """Return workbook row traceability data for the CU compliance JSON."""
    source = workbook_path or DEFAULT_WORKBOOK_PATH
    cases_by_cu = load_workbook_cases(source)
    tests_by_cu = tests_by_cu or {}
    exact_refs_by_row = exact_refs_by_row or {}

    by_cu: dict[str, dict[str, Any]] = {}
    total_cases = 0
    exact_case_count = 0
    for cu_key, cases in cases_by_cu.items():
        linked_nodeids = sorted(
            {str(test.get("nodeid", "")) for test in tests_by_cu.get(cu_key, []) if str(test.get("nodeid", "")).strip()}
        )
        case_json = []
        for case in cases:
            exact_nodeids = sorted(set(str(nodeid) for nodeid in exact_refs_by_row.get(case.ref, [])))
            if exact_nodeids:
                exact_case_count += 1
            case_json.append(case.to_json(linked_nodeids, exact_nodeids))
        total_cases += len(cases)
        by_cu[cu_key] = {
            "sheet": cases[0].sheet if cases else "",
            "case_count": len(cases),
            "positive_case_count": sum(1 for case in cases if case.case_type == "positive"),
            "negative_case_count": sum(1 for case in cases if case.case_type == "negative"),
            "linked_test_count": len(linked_nodeids),
            "cases": case_json,
        }

    return {
        "schema": "ijt-workbook-traceability/v1",
        "source": str(source.relative_to(PROJECT_ROOT)) if source.is_relative_to(PROJECT_ROOT) else str(source),
        "official_cu_count": len(cases_by_cu),
        "case_count": total_cases,
        "expected_case_count": EXPECTED_WORKBOOK_CASE_COUNT,
        "exact_case_count": exact_case_count,
        "mapping_precision": "exact" if exact_case_count == total_cases else "cu",
        "missing_case_cus": [cu_key for cu_key, cases in cases_by_cu.items() if not cases],
        "by_cu": by_cu,
    }
